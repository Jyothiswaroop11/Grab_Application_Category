import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from typing import List, Dict
import time
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE_URL = "https://applipedia.paloaltonetworks.com/"
CHROME_WAIT_TIME = 20  # Increased wait time


class ExcelHandler:
    def read_excel(self, file_path: str) -> List[str]:
        try:
            df = pd.read_excel(file_path)
            # Get the first column that contains URLs
            urls_column = df.iloc[:, 0]
            return urls_column.dropna().tolist()
        except Exception as e:
            raise Exception(f"Error reading Excel: {str(e)}")

    def write_excel(self, data: List[Dict], output_file: str) -> None:
        try:
            df = pd.DataFrame(data)
            writer = pd.ExcelWriter(output_file, engine='openpyxl')

            df[['SNO', 'Application', 'Category', 'Sub Category']].to_excel(
                writer,
                index=False,
                sheet_name='Sheet1'
            )

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Set column widths
            col_widths = {'A': 5, 'B': 20, 'C': 35, 'D': 35}
            for col, width in col_widths.items():
                worksheet.column_dimensions[col].width = width

            # Add colors and formatting
            header_fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Apply header formatting
            for cell in worksheet[1]:
                cell.fill = header_fill

            # Apply text wrapping, alignment, and conditional colors
            for row in worksheet.iter_rows(min_row=2, max_row=len(data) + 1):
                category_status = row[2].value  # Category column
                subcategory_status = row[3].value  # Sub Category column

                for idx, cell in enumerate(row):
                    cell.alignment = openpyxl.styles.Alignment(vertical='top', wrap_text=True)

                    # Color logic for SNO and Application columns (idx 0 and 1)
                    if idx in [0, 1]:
                        if "No Application Found" in [category_status, subcategory_status]:
                            cell.font = openpyxl.styles.Font(color="0000FF")  # Blue
                        elif "No Category Found" in [category_status, subcategory_status]:
                            cell.font = openpyxl.styles.Font(color="FF0000")  # Red
                        else:
                            cell.font = openpyxl.styles.Font(color="008000")  # Green

                    # Original color logic for Category and Sub Category columns
                    elif cell.value in ["No Category Found", "No Sub Category Found"]:
                        cell.font = openpyxl.styles.Font(color="FF0000")
                    elif cell.value in ["No Application Found"]:
                        cell.font = openpyxl.styles.Font(color="0000FF")
                    else:
                        cell.font = openpyxl.styles.Font(color="008000")

            writer.close()
        except Exception as e:
            raise Exception(f"Error writing Excel: {str(e)}")


class WebScraper:
    def __init__(self):
        self.driver = self._setup_driver()

    def _setup_driver(self) -> webdriver.Chrome:
        options = Options()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--headless")

        # SSL Certificate handling
        options.set_capability('acceptInsecureCerts', True)

        # Add proxy settings
        options.add_argument('--proxy-server=direct://')
        options.add_argument('--proxy-bypass-list=*')
        options.add_argument('--no-proxy-server')

        # Set environment variables for proxy
        os.environ['no_proxy'] = '*'
        os.environ['NO_PROXY'] = '*'
        os.environ['HTTP_PROXY'] = ''
        os.environ['HTTPS_PROXY'] = ''

        # Use ChromeDriverManager to handle driver installation
        service = Service(ChromeDriverManager().install())

        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(CHROME_WAIT_TIME)
        return driver

    def search_and_extract(self, url: str) -> Dict:
        try:
            self.driver.get(BASE_URL)
            WebDriverWait(self.driver, CHROME_WAIT_TIME).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )

            search_input = WebDriverWait(self.driver, CHROME_WAIT_TIME).until(
                EC.presence_of_element_located((By.ID, "tbSearch"))
            )
            search_button = WebDriverWait(self.driver, CHROME_WAIT_TIME).until(
                EC.element_to_be_clickable((By.ID, "btnSearch"))
            )

            search_input.clear()
            time.sleep(3)
            search_input.send_keys(url)
            search_button.click()
            time.sleep(5)

            # Check for invalid value alert
            try:
                alert = self.driver.switch_to.alert
                alert_text = alert.text
                if "Invalid value" in alert_text:
                    alert.accept()
                    return {
                        "Application": url,
                        "Category": "No Application Found",
                        "Sub Category": "No Application Found"
                    }
            except:
                pass

            categories = []
            subcategories = []

            category_elements = self.driver.find_elements(By.CSS_SELECTOR, "#CategoryList")
            for elem in category_elements:
                if elem.text.strip():
                    categories.append(elem.text.strip())

            subcategory_elements = self.driver.find_elements(By.CSS_SELECTOR, "#SubCategoryList")
            for elem in subcategory_elements:
                if elem.text.strip():
                    subcategories.append(elem.text.strip())

            if categories and subcategories:
                return {
                    "Application": url,
                    "Category": "\n".join(categories),
                    "Sub Category": "\n".join(subcategories)
                }

            raise Exception("No results found")

        except Exception as e:
            print(f"Error processing URL {url}: {str(e)}")
            return {
                "Application": url,
                "Category": "No Category Found",
                "Sub Category": "No Sub Category Found"
            }

    def close(self):
        if self.driver:
            self.driver.quit()


def process_urls(input_file: str, output_file: str) -> None:
    excel_handler = ExcelHandler()
    web_scraper = WebScraper()

    try:
        print("Reading input file...")
        urls = excel_handler.read_excel(input_file)
        results = []

        print(f"Processing {len(urls)} URLs...")
        for index, url in enumerate(urls, 1):
            print("=" * 150)
            print(f"Processing URL ::::::: {index}/{len(urls)} :::::::::::::::: {url}")
            result = web_scraper.search_and_extract(url)
            result['SNO'] = index
            results.append(result)
            time.sleep(2)  # Increased wait between requests

        print("Saving results...")
        excel_handler.write_excel(results, output_file)
        print(f"Results saved to {output_file}")

    except Exception as e:
        print(f"Error: {str(e)}")
    finally:
        web_scraper.close()


if __name__ == "__main__":
    input_file = "Excel_Files/inputfile.xlsx"
    output_file = "Excel_Files/output.xlsx"

    if not os.path.exists("Excel_Files"):
        os.makedirs("Excel_Files")

    process_urls(input_file, output_file)
